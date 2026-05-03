import requests
import xml.etree.ElementTree as ET
from dotenv import load_dotenv
from datetime import datetime
from dateutil.relativedelta import relativedelta
import os
import subprocess
import openpyxl
import google.generativeai as genai

load_dotenv()
api_key = os.getenv("PUBLIC_DATA_API_KEY")
gemini_api_key = os.getenv("GEMINI_API_KEY")

genai.configure(api_key=gemini_api_key)
model = genai.GenerativeModel("gemini-2.5-flash")

url = "https://apis.data.go.kr/1613000/RTMSDataSvcAptTrade/getRTMSDataSvcAptTrade"

districts = {
    "1": ("영등포구", "11560"),
    "2": ("관악구",   "11620"),
    "3": ("구로구",   "11530"),
    "4": ("강서구",   "11500")
}

# 지역 선택
print("=== 조회할 지역을 선택하세요 ===")
for key, (name, _) in districts.items():
    print(f"{key}. {name}")
print("5. 전체")

district_input = input("\n번호를 입력하세요: ")

# 년월 입력 및 유효성 검사
while True:
    year_month = input("조회할 년월을 입력하세요 (예: 202403): ")

    try:
        input_date = datetime.strptime(year_month, "%Y%m")
        today = datetime.today()
        this_month = today.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        min_date = this_month - relativedelta(years=10)
        last_month_str = (this_month - relativedelta(months=1)).strftime("%Y%m")

        if input_date >= this_month:
            print(f"이번 달 이후는 조회할 수 없어요. {last_month_str} 이전 날짜를 입력해주세요.\n")
        elif input_date < min_date:
            print(f"해당 날짜는 API 지원 범위를 벗어났어요. 10년 이내 날짜를 입력해주세요.\n")
        else:
            break

    except ValueError:
        print("날짜 형식이 올바르지 않아요. 예: 202403 형식으로 입력해주세요.\n")

# output 폴더 및 파일 경로 설정
output_dir = "output"
os.makedirs(output_dir, exist_ok=True)
file_path = os.path.join(output_dir, f"{year_month}.xlsx")

# 선택된 지역 추출
if district_input == "5":
    selected = list(districts.values())
else:
    selected = [districts[district_input]]

# 기존 파일 있으면 불러오고 없으면 새로 생성
if os.path.exists(file_path):
    wb = openpyxl.load_workbook(file_path)
else:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

for name, code in selected:

    # 이미 같은 구 시트가 있는지 확인
    if name in wb.sheetnames:
        print(f"\n{name} 데이터가 이미 {year_month}.xlsx 에 존재해요.")
        open_file = input("해당 파일을 열어드릴까요? [Y/N]: ").strip().upper()
        if open_file == "Y":
            subprocess.Popen(["start", "", file_path], shell=True)
        continue

    # API 호출
    params = {
        "serviceKey": api_key,
        "LAWD_CD": code,
        "DEAL_YMD": year_month,
        "numOfRows": "20"
    }

    response = requests.get(url, params=params)
    root = ET.fromstring(response.text)

    # 데이터 확인
    items = root.findall(".//item")
    if not items:
        print(f"\n{name} {year_month} 데이터가 없어요.")
        continue

    # 시트 생성 및 헤더 작성
    ws = wb.create_sheet(title=name)
    ws.append(["아파트명", "면적(㎡)", "거래금액(만원)", "층", "건축년도"])

    # 데이터 입력 및 AI 분석용 텍스트 준비
    data_for_ai = []
    for item in items:
        apt_name = item.findtext("aptNm")
        area     = item.findtext("excluUseAr")
        price    = item.findtext("dealAmount")
        floor    = item.findtext("floor")
        build_year = item.findtext("buildYear")
        ws.append([apt_name, area, price, floor, build_year])
        data_for_ai.append(f"아파트명: {apt_name}, 면적: {area}㎡, 거래금액: {price}만원, 층: {floor}층, 건축년도: {build_year}")

    # 컬럼 너비 자동 조정
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 4

    # Gemini AI 분석 요청
    print(f"{name} AI 분석 중...")
    data_text = "\n".join(data_for_ai)
    prompt = f"""
아래는 {year_month[:4]}년 {year_month[4:]}월 서울 {name} 아파트 실거래가 데이터예요.
이 데이터를 바탕으로 전반적인 시세 흐름을 3~5줄로 요약해주세요.
전문용어 없이 누구나 이해할 수 있게 작성해주세요.

{data_text}
"""
    ai_response = model.generate_content(prompt)
    ai_comment = ai_response.text

    # AI 분석 결과 시트에 추가
    ws.append([])
    ws.append(["AI 시세 분석"])
    ws.append([ai_comment])

    print(f"{name} 데이터 및 AI 분석 추가 완료!")

# 파일 저장
wb.save(file_path)
print(f"\nExcel 파일 저장 완료: {file_path}")